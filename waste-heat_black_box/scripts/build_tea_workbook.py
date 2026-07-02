#!/usr/bin/env python3
"""Build waste-heat SAWH black-box Excel TEA (mirrors waste-heat_lumped LCOW costing)."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_REPO = Path(__file__).resolve().parent.parent
_TEA_ROOT = _REPO.parent
_LUMPED_SRC = _TEA_ROOT / "waste-heat_lumped" / "src"
if str(_TEA_ROOT) not in sys.path:
    sys.path.insert(0, str(_TEA_ROOT))
if str(_LUMPED_SRC) not in sys.path:
    sys.path.insert(0, str(_LUMPED_SRC))

from tea_workbook_lib import compute_tea_metrics  # noqa: E402

from waste_heat_lumped.economics.bom import DEVICE_BOM_USD_PER_M2 as BOM  # noqa: E402
from waste_heat_lumped.economics.parasitic import default_electrical_loads  # noqa: E402
from waste_heat_lumped.economics.params import _load_economic_data  # noqa: E402
from waste_heat_lumped.physics.salt_properties import get_salt_price_usd_per_kg  # noqa: E402
from waste_heat_lumped.simulation.device_config import DeviceConfig  # noqa: E402
from waste_heat_lumped.simulation.ode_system import run_daily_operation  # noqa: E402
from waste_heat_lumped.weather.profiles import datacenter_baseline_profile  # noqa: E402

OUT = _REPO / "waste_heat_sawh_tea.xlsx"
DEFAULT_SALT = "LiCl"
DEFAULT_SALT_TO_POLYMER = 4.0


def _parasitic_load_inputs() -> tuple[tuple[str, float, float, float], ...]:
    return tuple(
        (
            load.name,
            load.shaft_power_w_per_m2,
            load.motor_efficiency,
            load.operating_hours_per_day,
        )
        for load in default_electrical_loads()
    )


def _baseline_simulation() -> tuple[float, float, int]:
    """Yield, thermal efficiency, and cycle count from datacenter-baseline daily sim."""
    config = DeviceConfig.datacenter_baseline()
    profile = datacenter_baseline_profile()
    yield_kg, eta, results = run_daily_operation(profile, config)
    return float(yield_kg), float(eta), len(results)


def _defaults() -> dict[str, float | str]:
    scalars, _ = _load_economic_data()
    daily_yield, thermal_efficiency, n_cycles = _baseline_simulation()
    return {
        "daily_yield_kg_per_m2": daily_yield,
        "cycles_per_day": 1.0,
        "mechanistic_cycles_per_day": float(n_cycles),
        "thermal_efficiency": thermal_efficiency,
        "salt_name": DEFAULT_SALT,
        "salt_to_polymer_ratio": DEFAULT_SALT_TO_POLYMER,
        "hydrogel_thickness_m": float(scalars["hydrogel_thickness_m"]),
        "salt_price_usd_per_kg": get_salt_price_usd_per_kg(DEFAULT_SALT),
        "hydrogel_density_kg_m3": float(scalars["hydrogel_density_kg_m3"]),
        "hydrogel_lifetime_years": float(scalars["hydrogel_lifetime_years"]),
        "c_acrylamide_usd_per_kg": float(scalars["c_acrylamide_usd_per_kg"]),
        "c_additives_usd_per_kg_composite": float(scalars["c_additives_usd_per_kg_composite"]),
        "discount_rate": float(scalars["discount_rate"]),
        "device_lifetime_years": int(scalars["device_lifetime_years"]),
        "total_investment_factor": float(scalars["total_investment_factor"]),
        "maintenance_cost_fraction": float(scalars["maintenance_cost_fraction"]),
        "utilization_factor": float(scalars["utilization_factor"]),
        "energy_cost_usd_per_year": float(scalars["energy_cost_usd_per_year"]),
        "energy_cost_usd_per_extra_half_cycle_per_day": float(
            scalars["energy_cost_usd_per_extra_half_cycle_per_day"]
        ),
        "electricity_price_usd_per_kwh": float(scalars["electricity_price_usd_per_kwh"]),
        "electric_heat_w_per_m2": 0.0,
        "desorption_hours_per_day": float(scalars["desorption_hours_per_day"]),
    }


INPUT_ROWS: tuple[tuple[str, str, str, str], ...] = (
    ("daily_yield_kg_per_m2", "Daily water yield", "kg/m²/d", "From waste-heat_lumped run_waste_heat_sim.py --profile datacenter-baseline --daily"),
    ("cycles_per_day", "Cycles per day", "1/d", "Set >1 only if daily yield is per cycle, not per day"),
    ("salt_name", "Salt", "—", "Catalog label (price entered below)"),
    ("salt_to_polymer_ratio", "Salt-to-polymer ratio", "—", "Mass ratio salt : polymer"),
    ("hydrogel_thickness_m", "Hydrogel thickness", "m", "Active layer thickness per m² footprint"),
    ("salt_price_usd_per_kg", "Salt price", "USD/kg", "LiCl default from salt catalog"),
    ("hydrogel_density_kg_m3", "Dry composite density", "kg/m³", ""),
    ("hydrogel_lifetime_years", "Hydrogel replacement interval", "yr", ""),
    ("c_acrylamide_usd_per_kg", "Binder / polymer cost", "USD/kg", ""),
    ("c_additives_usd_per_kg_composite", "Additives cost", "USD/kg composite", ""),
    ("discount_rate", "Discount rate", "—", "Real discount rate for CRF"),
    ("device_lifetime_years", "Device lifetime", "yr", ""),
    ("total_investment_factor", "Installed CAPEX multiplier", "—", "Erection / indirects on BOM (electrolyte_optimization)"),
    ("maintenance_cost_fraction", "Maintenance fraction", "—", "Annual O&M as fraction of installed CAPEX"),
    ("utilization_factor", "Utilization factor", "—", "Effective annual operating fraction"),
    ("energy_cost_usd_per_year", "Fixed energy cost", "USD/m²/yr", "Vacuum / auxiliary fixed energy (if allocated)"),
    ("energy_cost_usd_per_extra_half_cycle_per_day", "Extra cycle energy cost", "USD/m² per half-cycle above 1/d", ""),
    ("electricity_price_usd_per_kwh", "Electricity price", "USD/kWh", "For parasitic loads not covered by waste heat"),
    ("electric_heat_w_per_m2", "Supplemental electric heat", "W/m²", "0 when waste heat supplies desorption"),
    ("desorption_hours_per_day", "Desorption hours", "h/d", ""),
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
OUTPUT_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
BOLD = Font(bold=True)


def _style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _style_table(ws, row_start: int, row_end: int, col_end: int) -> None:
    for r in range(row_start, row_end + 1):
        for c in range(1, col_end + 1):
            ws.cell(row=r, column=c).border = BORDER


def _write_num(cell, value: float, *, fmt: str = "0.0000"):
    cell.value = float(value)
    cell.number_format = fmt
    return cell


def build() -> Path:
    defaults = _defaults()
    mechanistic_cycles = defaults.pop("mechanistic_cycles_per_day")
    thermal_efficiency = float(defaults.pop("thermal_efficiency"))
    parasitic_loads = _parasitic_load_inputs()
    load_specs = default_electrical_loads()
    metrics = compute_tea_metrics(
        defaults,
        BOM,
        electricity_label="Electricity (supplemental heat)",
        parasitic_loads=parasitic_loads,
    )

    wb = Workbook()
    ws_in = wb.active
    ws_in.title = "Inputs"

    ws_in["A1"] = "Waste-heat SAWH — black-box TEA inputs"
    ws_in["A1"].font = Font(bold=True, size=14)
    ws_in.merge_cells("A1:D1")
    ws_in["A2"] = (
        "Edit yellow cells, then re-run scripts/build_tea_workbook.py to refresh outputs. "
        "Hardware BOM from patent cost estimates; hydrogel OPEX and LCOW factors mirror "
        "electrolyte_optimization / waste-heat_lumped/economics/lcow.py"
    )
    ws_in.merge_cells("A2:D2")

    headers = ("Parameter", "Value", "Unit", "Notes")
    for col, h in enumerate(headers, start=1):
        ws_in.cell(row=4, column=col, value=h)
    _style_header_row(ws_in, 4, 4)

    for i, (key, label, unit, note) in enumerate(INPUT_ROWS):
        r = 5 + i
        ws_in.cell(row=r, column=1, value=label)
        val = defaults[key]
        c = ws_in.cell(row=r, column=2, value=val)
        c.fill = INPUT_FILL
        c.number_format = "General" if isinstance(val, str) else "0.0000"
        ws_in.cell(row=r, column=3, value=unit)
        ws_in.cell(row=r, column=4, value=note)

    _style_table(ws_in, 4, 4 + len(INPUT_ROWS), 4)
    ws_in.column_dimensions["A"].width = 34
    ws_in.column_dimensions["B"].width = 16
    ws_in.column_dimensions["C"].width = 14
    ws_in.column_dimensions["D"].width = 48

    r0 = 5 + len(INPUT_ROWS) + 2
    ws_in.cell(row=r0, column=1, value="Derived quantities").font = BOLD
    derived = (
        ("Mechanistic cycles per day (sim)", mechanistic_cycles, "1/d"),
        ("Thermal efficiency (vs Q_wh)", thermal_efficiency, "—"),
        ("Dry composite mass", metrics.dry_composite_mass_kg_m2, "kg/m²"),
        ("Capital recovery factor (CRF)", metrics.crf, "—"),
        ("Gross annual water", metrics.gross_annual_water_m3, "m³/m²/yr"),
        ("Net annual water", metrics.net_annual_water_m3, "m³/m²/yr"),
    )
    for j, (label, value, unit) in enumerate(derived):
        r = r0 + 1 + j
        ws_in.cell(row=r, column=1, value=label)
        c = ws_in.cell(row=r, column=2)
        _write_num(c, float(value))
        c.fill = OUTPUT_FILL
        ws_in.cell(row=r, column=3, value=unit)

    ws_c = wb.create_sheet("CAPEX")
    ws_c["A1"] = "Capital costs (USD per m² footprint)"
    ws_c["A1"].font = Font(bold=True, size=14)
    for col, h in enumerate(("Component", "Unit CAPEX", "Installed CAPEX", "Annualized CAPEX"), start=1):
        ws_c.cell(row=3, column=col, value=h)
    _style_header_row(ws_c, 3, 4)

    bom_start = 4
    for i, (name, cost) in enumerate(BOM):
        r = bom_start + i
        ws_c.cell(row=r, column=1, value=name)
        _write_num(ws_c.cell(row=r, column=2), cost, fmt="0.00")
        _write_num(ws_c.cell(row=r, column=3), metrics.installed_capex[i], fmt="0.00")
        _write_num(ws_c.cell(row=r, column=4), metrics.annualized_capex[i], fmt="0.00")
    bom_end = bom_start + len(BOM) - 1
    total_row = bom_end + 1
    ws_c.cell(row=total_row, column=1, value="Total device CAPEX").font = BOLD
    _write_num(ws_c.cell(row=total_row, column=2), sum(metrics.unit_capex), fmt="0.00").font = BOLD
    _write_num(ws_c.cell(row=total_row, column=3), metrics.installed_capex_total, fmt="0.00").font = BOLD
    _write_num(ws_c.cell(row=total_row, column=4), metrics.annualized_capex_total, fmt="0.00").font = BOLD
    _style_table(ws_c, 3, total_row, 4)
    for col in "ABCD":
        ws_c.column_dimensions[col].width = 22

    ws_e = wb.create_sheet("Parasitic electricity")
    ws_e["A1"] = "Parasitic grid electricity (USD per m² footprint per year)"
    ws_e["A1"].font = Font(bold=True, size=14)
    ws_e["A2"] = (
        "Grid power = shaft power / motor efficiency. "
        "Annual cost = electricity price × grid power × operating hours × 365 / 1000."
    )
    ws_e.merge_cells("A2:G2")
    elec_headers = (
        "Component",
        "Shaft power",
        "Motor efficiency",
        "Operating hours",
        "Grid power",
        "Annual kWh",
        "Annual USD",
        "Notes",
    )
    for col, h in enumerate(elec_headers, start=1):
        ws_e.cell(row=4, column=col, value=h)
    _style_header_row(ws_e, 4, len(elec_headers))
    elec_start = 5
    for i, (row, spec) in enumerate(zip(metrics.component_electricity, load_specs)):
        r = elec_start + i
        ws_e.cell(row=r, column=1, value=row.name)
        _write_num(ws_e.cell(row=r, column=2), row.shaft_power_w_per_m2, fmt="0.00")
        ws_e.cell(row=r, column=2).fill = INPUT_FILL
        _write_num(ws_e.cell(row=r, column=3), row.motor_efficiency, fmt="0.00")
        ws_e.cell(row=r, column=3).fill = INPUT_FILL
        _write_num(ws_e.cell(row=r, column=4), row.operating_hours_per_day, fmt="0.00")
        ws_e.cell(row=r, column=4).fill = INPUT_FILL
        _write_num(ws_e.cell(row=r, column=5), row.grid_power_w_per_m2, fmt="0.00")
        _write_num(ws_e.cell(row=r, column=6), row.annual_kwh_per_m2, fmt="0.00")
        _write_num(ws_e.cell(row=r, column=7), row.annual_cost_usd_per_m2, fmt="0.00")
        ws_e.cell(row=r, column=8, value=spec.notes)
    elec_end = elec_start + len(metrics.component_electricity) - 1
    total_elec_row = elec_end + 1
    ws_e.cell(row=total_elec_row, column=1, value="Total parasitic electricity").font = BOLD
    _write_num(
        ws_e.cell(row=total_elec_row, column=6),
        sum(row.annual_kwh_per_m2 for row in metrics.component_electricity),
        fmt="0.00",
    ).font = BOLD
    _write_num(
        ws_e.cell(row=total_elec_row, column=7),
        metrics.parasitic_electricity_annual,
        fmt="0.00",
    ).font = BOLD
    _style_table(ws_e, 4, total_elec_row, 8)
    for col in "ABCDEFGH":
        ws_e.column_dimensions[col].width = 18
    ws_e.column_dimensions["A"].width = 34
    ws_e.column_dimensions["H"].width = 42

    ws_o = wb.create_sheet("OPEX")
    ws_o["A1"] = "Operating costs (USD per m² footprint per year)"
    ws_o["A1"].font = Font(bold=True, size=14)
    for col, h in enumerate(("Cost item", "Type", "Annual USD/m²", "Notes"), start=1):
        ws_o.cell(row=3, column=col, value=h)
    _style_header_row(ws_o, 3, 4)

    opex_rows: list[tuple[str, str, float, str]] = [
        ("Maintenance", "Fixed", metrics.maintenance_annual, "Fraction of installed CAPEX"),
        ("Hydrogel replacement", "Fixed", metrics.hydrogel_annual, "Salt + polymer + additives amortized"),
        ("Fixed energy", "Fixed", metrics.fixed_energy_annual, "Vacuum/auxiliary if allocated"),
        ("Electricity (supplemental heat)", "Variable", metrics.electricity_annual, ""),
    ]
    for row in metrics.component_electricity:
        opex_rows.append(
            (
                f"Electricity: {row.name}",
                "Variable",
                row.annual_cost_usd_per_m2,
                "Shaft power / motor efficiency × hours",
            )
        )
    opex_rows.append(
        (
            "Extra cycling energy",
            "Variable",
            metrics.extra_cycle_energy_annual,
            "Energy above one half-cycle per day",
        )
    )
    opex_rows = tuple(opex_rows)
    opex_start = 4
    for i, (item, typ, amount, note) in enumerate(opex_rows):
        r = opex_start + i
        ws_o.cell(row=r, column=1, value=item)
        ws_o.cell(row=r, column=2, value=typ)
        _write_num(ws_o.cell(row=r, column=3), amount, fmt="0.00")
        ws_o.cell(row=r, column=4, value=note)
    opex_end = opex_start + len(opex_rows) - 1

    for r in range(opex_start, opex_end + 1):
        typ = ws_o.cell(row=r, column=2).value
        fill = PatternFill("solid", fgColor="DDEBF7") if typ == "Fixed" else PatternFill("solid", fgColor="FCE4D6")
        ws_o.cell(row=r, column=2).fill = fill

    fix_row = opex_end + 2
    var_row = opex_end + 3
    tot_row = opex_end + 4
    ws_o.cell(row=fix_row, column=1, value="Subtotal — fixed OPEX").font = BOLD
    _write_num(ws_o.cell(row=fix_row, column=3), metrics.fixed_opex_total, fmt="0.00")
    ws_o.cell(row=var_row, column=1, value="Subtotal — variable OPEX").font = BOLD
    _write_num(ws_o.cell(row=var_row, column=3), metrics.variable_opex_total, fmt="0.00")
    ws_o.cell(row=tot_row, column=1, value="Total OPEX").font = BOLD
    _write_num(ws_o.cell(row=tot_row, column=3), metrics.total_opex, fmt="0.00").font = BOLD
    _style_table(ws_o, 3, tot_row, 4)
    for col in "ABCD":
        ws_o.column_dimensions[col].width = 26

    ws_s = wb.create_sheet("LCOW")
    ws_s["A1"] = "Levelized cost of water"
    ws_s["A1"].font = Font(bold=True, size=14)
    ws_s["A2"] = "LCOW = (annualized CAPEX + total OPEX) / net annual water production"
    ws_s.merge_cells("A2:D2")

    summary: tuple[tuple[str, float, str], ...] = (
        ("Annualized CAPEX", metrics.annualized_capex_total, "USD/m²/yr"),
        ("Total fixed OPEX", metrics.fixed_opex_total, "USD/m²/yr"),
        ("Total variable OPEX", metrics.variable_opex_total, "USD/m²/yr"),
        ("Total annual cost", metrics.total_annual_cost, "USD/m²/yr"),
        ("Gross annual water", metrics.gross_annual_water_m3, "m³/m²/yr"),
        ("Net annual water", metrics.net_annual_water_m3, "m³/m²/yr"),
    )
    for col, h in enumerate(("Metric", "Value", "Unit"), start=1):
        ws_s.cell(row=4, column=col, value=h)
    _style_header_row(ws_s, 4, 3)
    for i, (label, value, unit) in enumerate(summary):
        r = 5 + i
        ws_s.cell(row=r, column=1, value=label)
        _write_num(ws_s.cell(row=r, column=2), value)
        ws_s.cell(row=r, column=3, value=unit)

    lcow_row = 5 + len(summary) + 1
    ws_s.cell(row=lcow_row, column=1, value="LCOW").font = Font(bold=True, size=12)
    lcow_cell = ws_s.cell(row=lcow_row, column=2)
    _write_num(lcow_cell, metrics.lcow_usd_per_m3, fmt="0.0000")
    lcow_cell.font = Font(bold=True, size=12)
    lcow_cell.fill = OUTPUT_FILL
    ws_s.cell(row=lcow_row, column=3, value="USD/m³").font = BOLD

    br = lcow_row + 3
    ws_s.cell(row=br, column=1, value="LCOW cost breakdown").font = BOLD
    ws_s.merge_cells(start_row=br, start_column=1, end_row=br, end_column=3)
    br += 1
    for col, h in enumerate(("Segment", "Annual USD/m²", "USD/m³ water"), start=1):
        ws_s.cell(row=br, column=col, value=h)
    _style_header_row(ws_s, br, 3)
    br += 1
    for label, annual_usd, usd_per_m3 in zip(
        [x[0] for x in metrics.lcow_breakdown_annual],
        [x[1] for x in metrics.lcow_breakdown_annual],
        [x[1] for x in metrics.lcow_breakdown_usd_per_m3],
    ):
        ws_s.cell(row=br, column=1, value=label)
        _write_num(ws_s.cell(row=br, column=2), annual_usd, fmt="0.0000")
        _write_num(ws_s.cell(row=br, column=3), usd_per_m3, fmt="0.0000")
        br += 1
    ws_s.cell(row=br, column=1, value="Total (check)").font = BOLD
    _write_num(
        ws_s.cell(row=br, column=3),
        sum(x[1] for x in metrics.lcow_breakdown_usd_per_m3),
        fmt="0.0000",
    ).font = BOLD

    _style_table(ws_s, 4, br, 3)
    ws_s.column_dimensions["A"].width = 28
    ws_s.column_dimensions["B"].width = 18
    ws_s.column_dimensions["C"].width = 18

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
