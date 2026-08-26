"""Loads the Physics, Economics and Salts parameter tables directly from
``docs/parameters.xlsx`` -- the repo-wide single source of truth for system physics
constants, LCOW/NPV economics inputs, and salt properties. ``waste_heat`` and
``sawh_bayesopt`` read the same workbook through this module; ``waste_heat``'s rows
carry a "Waste-heat" prefix where its value differs from the solar device's.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

_XLSX_PATH = Path(__file__).resolve().parents[2] / "docs" / "parameters.xlsx"


def _load_sheet(sheet_name: str) -> dict[str, dict[str, Any]]:
    """Columns are located by header, not by position: the sheets are hand-edited in
    Excel, and a positional unpack turns "someone inserted a column" into a ValueError
    at import time that takes every package down with it."""
    wb = openpyxl.load_workbook(_XLSX_PATH, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = ["" if h is None else str(h).strip() for h in next(rows)]
    col = {
        key: header.index(key)
        for key in ("Name", "Value", "Lower (for Sweeps)", "Upper (for Sweeps)", "Source")
    }
    return {
        str(row[col["Name"]]).strip(): {
            "value": row[col["Value"]],
            "lower": row[col["Lower (for Sweeps)"]],
            "upper": row[col["Upper (for Sweeps)"]],
            "source": row[col["Source"]],
        }
        for row in rows
        if row[col["Name"]] is not None
    }


def _load_salts() -> dict[str, dict[str, Any]]:
    """The Salts sheet, keyed by salt name.

    Its shape is one row per salt rather than Physics/Economics' one row per named
    scalar, so it gets its own reader: a 4-salt x 9-property table cannot be expressed
    in a Name -> Value sheet without splitting the salts apart.
    """
    wb = openpyxl.load_workbook(_XLSX_PATH, data_only=True, read_only=True)
    ws = wb["Salts"]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows)]
    return {
        str(row[0]).strip(): dict(zip(header[1:], row[1:]))
        for row in rows
        if row[0] is not None
    }


PHYSICS: dict[str, dict[str, Any]] = _load_sheet("Physics")
ECONOMICS: dict[str, dict[str, Any]] = _load_sheet("Economics")
SALTS: dict[str, dict[str, Any]] = _load_salts()


def physics_value(name: str, *, mm_to_m: bool = False) -> float:
    value = float(PHYSICS[name]["value"])
    return value / 1000.0 if mm_to_m else value


def physics_bounds(name: str, *, mm_to_m: bool = False) -> tuple[float, float]:
    """The row's ``Lower (for Sweeps)`` / ``Upper (for Sweeps)`` pair."""
    row = PHYSICS[name]
    lower, upper = float(row["lower"]), float(row["upper"])
    return (lower / 1000.0, upper / 1000.0) if mm_to_m else (lower, upper)


def economics_value(name: str) -> Any:
    return ECONOMICS[name]["value"]
