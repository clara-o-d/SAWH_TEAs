"""Read LCOW breakdown from black-box TEA workbooks and plot stacked bar charts."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Patch
from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class LcowBreakdown:
    title: str
    lcow_usd_per_m3: float
    segments: tuple[tuple[str, float, float], ...]  # label, annual USD/m², USD/m³


def read_lcow_breakdown(workbook_path: Path | str) -> LcowBreakdown:
    """Parse the LCOW cost breakdown table from a black-box TEA workbook."""
    path = Path(workbook_path)
    wb = load_workbook(path, data_only=True)
    ws = wb["LCOW"]

    title = str(ws["A1"].value or "LCOW breakdown")
    lcow_usd_per_m3 = 0.0
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=2, values_only=True):
        if row[0] == "LCOW" and row[1] is not None:
            lcow_usd_per_m3 = float(row[1])
            break

    header_row: int | None = None
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Segment":
            header_row = row_idx
            break
    if header_row is None:
        raise ValueError(f"No LCOW cost breakdown table found in {path}")

    segments: list[tuple[str, float, float]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str or label_str.lower().startswith("total"):
            break
        annual = ws.cell(row=row_idx, column=2).value
        usd_per_m3 = ws.cell(row=row_idx, column=3).value
        if annual is None or usd_per_m3 is None:
            continue
        segments.append((label_str, float(annual), float(usd_per_m3)))

    if not segments:
        raise ValueError(f"LCOW breakdown table is empty in {path}")

    return LcowBreakdown(
        title=title,
        lcow_usd_per_m3=lcow_usd_per_m3,
        segments=tuple(segments),
    )


def _segment_colors(labels: list[str]) -> list[str]:
    capex_labels = [label for label in labels if label.startswith("CAPEX:")]
    capex_colors = plt.cm.Blues(np.linspace(0.45, 0.9, max(len(capex_labels), 1)))
    capex_map = dict(zip(capex_labels, capex_colors, strict=False))

    fixed = {
        "Maintenance": "#8172B3",
        "Hydrogel: salt": "#55A868",
        "Hydrogel: acrylamide": "#88C999",
        "Hydrogel: additives": "#AAD4B8",
        "Fixed energy": "#C44E52",
        "Extra cycling energy": "#CCB974",
    }
    colors: list[str] = []
    elec_idx = 0
    elec_cmap = plt.cm.Oranges(np.linspace(0.45, 0.85, 8))
    for label in labels:
        if label in fixed:
            colors.append(fixed[label])
        elif label.startswith("CAPEX:"):
            colors.append(capex_map[label])
        elif label.startswith("Electricity"):
            colors.append(elec_cmap[min(elec_idx, len(elec_cmap) - 1)])
            elec_idx += 1
        else:
            colors.append("#777777")
    return colors


def _text_color(color) -> str:
    if isinstance(color, str):
        hex_color = color.lstrip("#")
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    else:
        r, g, b = (int(c * 255) for c in color[:3])
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return "white" if luminance < 140 else "black"


def plot_lcow_breakdown_stacked(
    breakdown: LcowBreakdown,
    *,
    output_path: Path | str,
    min_usd_per_m3: float = 1e-6,
) -> Path:
    """Write a vertical stacked bar chart of LCOW segments (USD/m³)."""
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    segments = [
        (label, annual, usd_per_m3)
        for label, annual, usd_per_m3 in breakdown.segments
        if usd_per_m3 > min_usd_per_m3
    ]
    if not segments:
        raise ValueError("No non-zero LCOW segments to plot")

    labels = [label for label, _, _ in segments]
    values = [usd_per_m3 for _, _, usd_per_m3 in segments]
    colors = _segment_colors(labels)
    total = sum(values)

    fig_w = max(8.0, 0.22 * len(segments) + 6.5)
    fig, (ax_bar, ax_leg) = plt.subplots(
        1,
        2,
        figsize=(fig_w, 6),
        gridspec_kw={"width_ratios": [1.4, 1.6]},
    )

    bottom = 0.0
    for label, value, color in zip(labels, values, colors, strict=True):
        ax_bar.bar(
            0,
            value,
            bottom=bottom,
            width=0.55,
            color=color,
            edgecolor="white",
            linewidth=0.6,
        )
        if value / total >= 0.06:
            ax_bar.text(
                0,
                bottom + value / 2,
                f"{value:.1f}",
                ha="center",
                va="center",
                fontsize=8,
                color="white" if _text_color(color) == "white" else "black",
                fontweight="bold",
            )
        bottom += value

    ax_bar.set_ylim(0, bottom * 1.02)
    ax_bar.set_xticks([])
    ax_bar.set_ylabel("LCOW (USD/m³)")
    ax_bar.set_title(
        f"{breakdown.title}\nTotal LCOW = ${breakdown.lcow_usd_per_m3:.2f}/m³",
        loc="left",
        fontsize=11,
    )
    ax_bar.grid(axis="y", color="#E0E0E0", linewidth=0.8)
    ax_bar.set_axisbelow(True)

    ax_leg.axis("off")
    handles = [
        Patch(facecolor=color, edgecolor="white", label=f"{label} ({value:.2f})")
        for label, value, color in zip(labels, values, colors, strict=True)
    ]
    ax_leg.legend(
        handles=handles,
        loc="upper left",
        frameon=False,
        fontsize=8,
        title="Cost segments",
        title_fontsize=9,
    )

    fig.tight_layout()
    fig.savefig(out, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return out


def plot_lcow_breakdown_from_workbook(
    workbook_path: Path | str,
    *,
    output_path: Path | str,
) -> tuple[LcowBreakdown, Path]:
    breakdown = read_lcow_breakdown(workbook_path)
    out = plot_lcow_breakdown_stacked(breakdown, output_path=output_path)
    return breakdown, out
